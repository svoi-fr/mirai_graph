from neo4j import GraphDatabase
import openpyxl
import json
from FlagEmbedding import BGEM3FlagModel
import os

EXCEL_FILE_PATH = "output.xlsx"
NEO4J_URI = os.environ.get("NEO4J_URI", "bolt://localhost:7687")  # Get from environment or default
NEO4J_USER = os.environ.get("NEO4J_USER", "neo4j")  # Get from environment
NEO4J_PASSWORD = os.environ.get("NEO4J_PASSWORD")  # Get from environment
BATCH_SIZE = 100  # Adjust the batch size as needed

# Initialize the BAAI BGE M3 embedding model
model = BGEM3FlagModel('BAAI/bge-m3', use_fp16=True)

def create_neo4j_nodes(excel_file_path, neo4j_uri, neo4j_user, neo4j_password):
    """Creates Neo4j nodes from Excel data without relationships."""
    driver = GraphDatabase.driver(neo4j_uri, auth=(neo4j_user, neo4j_password))
    wb = openpyxl.load_workbook(excel_file_path)
    ws_data = wb["LLM Responses"]

    with driver.session() as session:
        for row in ws_data.iter_rows(min_row=2):  # Skip header row
            url, _, response_cell = row
            url = url.value
            response_json = json.loads(response_cell.value)

            # --- Create Document node and embedding ---
            summary = response_json["Summary"]
            document_embedding = model.encode([summary.get("info", "")], max_length=8192)['dense_vecs'][0].tolist()
            session.run(
                """
                MERGE (d:Document {id: $summary_id})
                SET d.country = $country,
                    d.info = $description,
                    d.type = $type,
                    d.url = $url,
                    d.embedding = $embedding
                """,
                summary_id=summary["id"],
                country=summary.get("country", ""),
                description=summary.get("info", ""),
                type=summary["type"],
                url=url,
                embedding=document_embedding
            )

            # --- Create nodes without relationships ---
            for node_type in ["Context", "Provision", "Task", "Organization", "Contact"]:
                nodes = response_json.get(node_type, [])
                if nodes:
                    info_texts = [node.get("info", "") for node in nodes]
                    embeddings = model.encode(info_texts, max_length=8192)['dense_vecs']
                    
                    for node, embedding in zip(nodes, embeddings):
                        node_id = node["id"]
                        node_label = node_type

                        # --- Merge node (create or update) ---
                        session.run(
                            f"""
                            MERGE (n:{node_label} {{id: $node_id}})
                            ON CREATE SET n.info = $info, n.embedding = $embedding
                            ON MATCH SET 
                                n.info = CASE WHEN $info <> '' THEN coalesce(n.info, $info) ELSE n.info END,
                                n.url = CASE WHEN $url <> '' THEN coalesce(n.url, $url) ELSE n.url END,
                                n.phone = CASE WHEN $phone <> '' THEN coalesce(n.phone, $phone) ELSE n.phone END,
                                n.email = CASE WHEN $email <> '' THEN coalesce(n.email, $email) ELSE n.email END,
                                n.address = CASE WHEN $address <> '' THEN coalesce(n.address, $address) ELSE n.address END,
                                n.city = CASE WHEN $city <> '' THEN coalesce(n.city, $city) ELSE n.city END,
                                n.embedding = CASE WHEN $embedding <> [] THEN coalesce(n.embedding, $embedding) ELSE n.embedding END
                            WITH n, $cities AS new_cities
                            UNWIND new_cities AS new_city
                            MERGE (n)-[:IN_CITY]->(city:City {{name: new_city}})
                            """,
                            node_id=node_id,
                            info=node.get("info", ""),
                            url=node.get("url", ""),
                            phone=node.get("phone", ""),
                            email=node.get("email", ""),
                            address=node.get("address", ""),
                            city=node.get("city", ""),
                            cities=node.get("cities", []),
                            embedding=embedding.tolist()
                        )

                        # --- Connect node to the document ---
                        session.run(
                            """
                            MATCH (n {id: $node_id}), (d:Document {id: $summary_id})
                            MERGE (d)-[:DESCRIBE]->(n)
                            """,
                            node_id=node_id,
                            summary_id=summary["id"]
                        )

    driver.close()
    print("Neo4j nodes created successfully!")

def create_neo4j_relationships(excel_file_path, neo4j_uri, neo4j_user, neo4j_password):
    """Creates relationships between existing Neo4j nodes based on Excel data."""
    driver = GraphDatabase.driver(neo4j_uri, auth=(neo4j_user, neo4j_password))
    wb = openpyxl.load_workbook(excel_file_path)
    ws_data = wb["LLM Responses"]

    with driver.session() as session:
        for row in ws_data.iter_rows(min_row=2):  # Skip header row
            _, _, response_cell = row
            response_json = json.loads(response_cell.value)

            # --- Create relationships between nodes ---
            for node_type in ["Context", "Provision", "Task", "Organization", "Contact"]:
                nodes = response_json.get(node_type, [])
                for node in nodes:
                    node_id = node["id"]
                    node_label = node_type

                    # --- Create customized relationships ---
                    if node_label == "Task":
                        # Tasks APPLY to Contexts
                        for ref_id in node.get("contexts", []):
                            session.run(
                                """
                                MATCH (task:Task {id: $node_id}), (context:Context {id: $ref_id})
                                MERGE (task)-[:APPLY]->(context)
                                """,
                                node_id=node_id,
                                ref_id=ref_id
                            )

                    if node_label == "Provision":
                        # Provisions ADDRESS Contexts
                        for ref_id in node.get("contexts", []):
                            session.run(
                                """
                                MATCH (provision:Provision {id: $node_id}), (context:Context {id: $ref_id})
                                MERGE (provision)-[:ADDRESS]->(context)
                                """,
                                node_id=node_id,
                                ref_id=ref_id
                            )
                        
                        # Organizations PROVIDE Provisions
                        for ref_id in node.get("organizations", []):
                            session.run(
                                """
                                MATCH (provision:Provision {id: $node_id}), (org:Organization {id: $ref_id})
                                MERGE (org)-[:PROVIDES]->(provision)
                                """,
                                node_id=node_id,
                                ref_id=ref_id
                            )

                    if node_label == "Organization":
                        # Organizations CONTACT Contacts
                        for ref_id in node.get("contacts", []):
                            session.run(
                                """
                                MATCH (org:Organization {id: $node_id}), (contact:Contact {id: $ref_id})
                                MERGE (org)-[:CONTACT]->(contact)
                                """,
                                node_id=node_id,
                                ref_id=ref_id
                            )

    driver.close()
    print("Neo4j relationships created successfully!")

if __name__ == "__main__":
    # Phase 1: Create all nodes
    create_neo4j_nodes(EXCEL_FILE_PATH, NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD)
    
    # Phase 2: Create relationships between nodes
    create_neo4j_relationships(EXCEL_FILE_PATH, NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD)
