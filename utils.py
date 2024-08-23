import openpyxl
import json
import random

EXCEL_FILE_PATH = "output.xlsx"
PROCESSED_URLS_SHEET = "Processed URLs"

def append_to_excel(url, original_text, response, excel_file_path=EXCEL_FILE_PATH):
    try:
        wb = openpyxl.load_workbook(excel_file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "LLM Responses"
        ws_data.append(["URL", "Original Text", "LLM Response"]) 

    if "LLM Responses" in wb.sheetnames:
        ws_data = wb["LLM Responses"]
    else:
        ws_data = wb.active
        ws_data.title = "LLM Responses"
        ws_data.append(["URL", "Original Text", "LLM Response"]) 

    ws_data.append([url.encode('utf-8'), 
                    original_text.encode('utf-8'), 
                    response.encode('utf-8')])

    ws_data.column_dimensions['A'].width = 50
    ws_data.column_dimensions['B'].width = 100
    ws_data.column_dimensions['C'].width = 100

    if PROCESSED_URLS_SHEET in wb.sheetnames:
        ws_urls = wb[PROCESSED_URLS_SHEET]
    else:
        ws_urls = wb.create_sheet(PROCESSED_URLS_SHEET)

    ws_urls.append([url])

    wb.save(excel_file_path)
    print(f"Data for {url} saved to {excel_file_path}")

def load_processed_urls(excel_file_path=EXCEL_FILE_PATH):
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        if PROCESSED_URLS_SHEET in wb.sheetnames:
            ws_urls = wb[PROCESSED_URLS_SHEET]
            processed_urls = set(cell.value for row in ws_urls.iter_rows() for cell in row if cell.value)
            return processed_urls
        else:
            return set()
    except FileNotFoundError:
        return set()


def generate_jsonl_from_excel(excel_file_path, jsonl_file_path, mistral_training_path="mistral_training.jsonl", mistral_validation_path="mistral_validation.jsonl"):
    """Generates the JSONL files from the Excel data."""
    wb = openpyxl.load_workbook(excel_file_path)
    ws_data = wb["LLM Responses"]
    data = []

    for row in ws_data.iter_rows(min_row=2):
        url, original_text, response = row
        json_data = {
            "messages": [
                {"role": "user", "content": original_text.value},
                {"role": "assistant", "content": response.value}
            ]
        }
        data.append(json_data)

    # Shuffle data for random split
    random.shuffle(data)

    # 80/20 training/validation split
    split_index = int(0.8 * len(data))
    training_data = data[:split_index]
    validation_data = data[split_index:]

    # Write to output.jsonl
    with open(jsonl_file_path, 'w', encoding='utf-8') as f:
        for item in data:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")
    print(f"JSONL file generated from {excel_file_path} to {jsonl_file_path}")

    # Write to mistral_training.jsonl
    with open(mistral_training_path, 'w', encoding='utf-8') as f:
        for item in training_data:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")
    print(f"Mistral training JSONL file generated from {excel_file_path} to {mistral_training_path}")

    # Write to mistral_validation.jsonl
    with open(mistral_validation_path, 'w', encoding='utf-8') as f:
        for item in validation_data:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")
    print(f"Mistral validation JSONL file generated from {excel_file_path} to {mistral_validation_path}")